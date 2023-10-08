import telebot
import random
import openpyxl

# Crear el bot con el token de acceso
TOKEN='6636288744:AAFhyH1hnrilP2fBGutv4F_WxSy8OV2XtCc'
bot = telebot.TeleBot(TOKEN)



# Crea un nuevo archivo Excel y agrega los encabezados de columna
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Usuarios"
worksheet.cell(row=1, column=1, value="ID")
worksheet.cell(row=1, column=2, value="Nombre de usuario")
worksheet.cell(row=1, column=3, value="Número generado")

# Crea un diccionario para almacenar los números generados para cada usuario
numeros_generados = {}

@bot.message_handler(commands=['start'])

def send_welcome(message):
 # Enviar mensaje de bienvenida con botones
    bot.reply_to(message, f"👋 Hola @{message.from_user.username}!\n\n❗️IMPORTANTE:\n\n🧏 Para recibir el premio (en caso de que quede como ganador) tiene que estar subscrito al canal @fifacuba de no estarlo al momento que se anuncie el ganador, se repite el sorteo a una persona que si este subscrita. Asi que como requisitos fundamental para participar es:\n\n👌 ESTAR SUBSCRITO AL CANAL @fifacuba\n\n🫡 Presione /sorteo para generar su # para el sorteo de @fifacuba.\n\n🎱 Presione /numero para su recordar su número\n\n👮 Creado por @Alex_GlezRM")
 #Manejar el comando /archivo1

# Define una función que genera un número aleatorio para cada usuario que ingresa al bot
@bot.message_handler(commands=['sorteo'])
def generar_numero(message):
    chat_id = message.chat.id
    if chat_id not in numeros_generados:
        numero_generado = random.randint(1, 50)
        numeros_generados[chat_id] = numero_generado
        bot.reply_to(message, f"🍀 Tu número generado es {numero_generado}\n\n👮 Creado por @Alex_GlezRM")
        # Agrega la información del usuario al archivo Excel
        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=chat_id)
        worksheet.cell(row=row, column=2, value=message.chat.username)
        worksheet.cell(row=row, column=3, value=numero_generado)
        workbook.save("usuarios.xlsx")
    else:
        bot.reply_to(message, f"🍀 Ya tiene número para el sorteo, especificamente el: {numeros_generados[chat_id]} ¡No intente sacar otro!\n\n👮 Creado por @Alex_GlezRM")

# Define una función que envía un mensaje a todos los usuarios cuando se han asignado los 100 números
def enviar_mensaje_lista_llena():
    if len(numeros_generados) == 100:
        bot.send_message("🍀 La lista está llena, no hay más números disponibles\n\n👮 Creado por @Alex_GlezRM")

# Define una función que verifica si un usuario ya ha sido asignado un número
@bot.message_handler(commands=['numero'])
def verificar_numero(message):
    chat_id = message.chat.id
    if chat_id in numeros_generados:
        bot.reply_to(message, f"🍀 Tu número generado es {numeros_generados[chat_id]}\n\n👮 Creado por @Alex_GlezRM")
    else:
        bot.reply_to(message, "🍀 Aún no tienes un número generado, generelo presionando /sorteo\n\n👮 Creado por @Alex_GlezRM")

bot.add_message_handler(enviar_mensaje_lista_llena)

bot.polling()
